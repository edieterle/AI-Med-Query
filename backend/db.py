from datetime import timedelta
import random
import pandas as pd
from openpyxl import load_workbook
from faker import Faker
from sqlalchemy import (
    create_engine, Column, Integer, String, DECIMAL,
    Date, ForeignKey, TIMESTAMP, Boolean
)
from sqlalchemy.orm import sessionmaker, declarative_base, relationship

# -------------------------
# Database setup
# -------------------------
DATABASE_URL = "postgresql+psycopg2://postgres:1774@localhost:5432/postgres"
engine = create_engine(DATABASE_URL, echo=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# -------------------------
# Table definitions
# -------------------------
class Patient(Base):
    __tablename__ = "patients"
    patient_id = Column(Integer, primary_key=True, index=True)
    age = Column(Integer)
    sex = Column(String(10))
    weight = Column(DECIMAL(5, 2))
    height = Column(DECIMAL(5, 2))
    diagnosis_code = Column(String(10))
    devices = relationship("Device", back_populates="patient")

class Device(Base):
    __tablename__ = "devices"
    device_id = Column(Integer, primary_key=True, index=True)
    patient_id = Column(Integer, ForeignKey("patients.patient_id"))
    device_type = Column(String(50))
    implant_date = Column(Date)
    manufacturer = Column(String(50))
    patient = relationship("Patient", back_populates="devices")
    readings = relationship("Reading", back_populates="device")
    outcomes = relationship("Outcome", back_populates="device")

class Reading(Base):
    __tablename__ = "readings"
    reading_id = Column(Integer, primary_key=True, index=True)
    device_id = Column(Integer, ForeignKey("devices.device_id"))
    timestamp = Column(TIMESTAMP)
    heart_rate = Column(Integer)
    blood_pressure = Column(String(7))
    battery_level = Column(Integer)
    device = relationship("Device", back_populates="readings")

class Outcome(Base):
    __tablename__ = "outcomes"
    outcome_id = Column(Integer, primary_key=True, index=True)
    device_id = Column(Integer, ForeignKey("devices.device_id"))
    complication_occurred = Column(Boolean)
    device_replacement_needed = Column(Boolean)
    time_to_failure = Column(Integer)
    readmission_within_30_days = Column(Boolean)
    device = relationship("Device", back_populates="outcomes")

# -------------------------
# Utility functions
# -------------------------
def init_db():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)

def drop_db():
    Base.metadata.drop_all(bind=engine)

# -------------------------
# Data generation helpers
# -------------------------
fake = Faker()

def generate_patient():
    age = random.randint(20, 90)
    sex = random.choices(["Male", "Female"], weights=[0.48, 0.52])[0]
    height = round(random.uniform(150, 200), 2)
    weight = round(random.uniform(50, 120), 2)
    diagnosis_code = f"D{random.randint(100, 999)}"

    # Risk score for complications
    risk_score = 0
    if age > 65:
        risk_score += 1.5
    if sex == "Male":
        risk_score += 0.5
    bmi = weight / ((height/100)**2)
    if bmi > 30:
        risk_score += 1

    return {
        "age": age,
        "sex": sex,
        "weight": weight,
        "height": height,
        "diagnosis_code": diagnosis_code,
        "risk_score": risk_score
    }

def generate_reading(risk_score):
    heart_rate = random.randint(60, 80) + int(risk_score * 5)
    heart_rate = min(heart_rate, 130)
    systolic = random.randint(110, 120) + int(risk_score * 5)
    diastolic = random.randint(70, 80) + int(risk_score * 3)
    systolic = min(systolic, 180)
    diastolic = min(diastolic, 120)
    blood_pressure = f"{systolic}/{diastolic}"
    battery_level = random.randint(50, 100)
    return {
        "heart_rate": heart_rate,
        "blood_pressure": blood_pressure,
        "battery_level": battery_level
    }

def generate_reading_over_time(risk_score, day):
    reading = generate_reading(risk_score)
    reading["heart_rate"] += int(day * 0.2 * risk_score)   # gradual trend
    reading["battery_level"] = max(reading["battery_level"] - day, 0)
    return reading

def generate_outcome(device_type, risk_score):
    base_prob = 0.02
    if device_type == "Pacemaker":
        base_prob += 0.05
    elif device_type == "Defibrillator":
        base_prob += 0.03

    prob_complication = base_prob + 0.1 * risk_score

    complication_occurred = random.random() < prob_complication
    device_replacement_needed = complication_occurred and random.random() < 0.6
    readmission_within_30_days = complication_occurred and random.random() < 0.4
    time_to_failure = random.randint(30, 1000) if complication_occurred else random.randint(1000, 2000)

    return {
        "complication_occurred": complication_occurred,
        "device_replacement_needed": device_replacement_needed,
        "readmission_within_30_days": readmission_within_30_days,
        "time_to_failure": time_to_failure
    }

# -------------------------
# Main generator
# -------------------------
def generate_data(n_patients=20):
    init_db()
    db = SessionLocal()

    device_types = ["Pacemaker", "Defibrillator", "Neurostimulator"]
    manufacturers = ["Medtronic", "Boston Scientific", "Abbott"]

    patients = []
    for _ in range(n_patients):
        p_data = generate_patient()
        patient = Patient(
            age=p_data["age"],
            sex=p_data["sex"],
            weight=p_data["weight"],
            height=p_data["height"],
            diagnosis_code=p_data["diagnosis_code"]
        )
        db.add(patient)
        db.commit()
        db.refresh(patient)
        p_data["db_obj"] = patient
        patients.append(p_data)

        # 1-3 devices per patient
        for _ in range(random.randint(1, 3)):
            device_type = random.choice(device_types)
            implant_date = fake.date_between(start_date='-5y', end_date='today')
            device = Device(
                patient_id=patient.patient_id,
                device_type=device_type,
                implant_date=implant_date,
                manufacturer=random.choice(manufacturers)
            )
            db.add(device)
            db.commit()
            db.refresh(device)

            # 30 days of readings per device
            for day in range(30):
                timestamp = implant_date + timedelta(days=day)
                r_data = generate_reading_over_time(p_data["risk_score"], day)
                reading = Reading(
                    device_id=device.device_id,
                    timestamp=timestamp,
                    heart_rate=r_data["heart_rate"],
                    blood_pressure=r_data["blood_pressure"],
                    battery_level=r_data["battery_level"]
                )
                db.add(reading)

            # Outcome per device
            o_data = generate_outcome(device_type, p_data["risk_score"])
            outcome = Outcome(
                device_id=device.device_id,
                complication_occurred=o_data["complication_occurred"],
                device_replacement_needed=o_data["device_replacement_needed"],
                readmission_within_30_days=o_data["readmission_within_30_days"],
                time_to_failure=o_data["time_to_failure"]
            )
            db.add(outcome)

    db.commit()
    db.close()

# -------------------------
# Export function
# -------------------------
def export_to_excel(output_file="exported_data.xlsx"):
    tables = ["patients", "devices", "readings", "outcomes"]
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for table in tables:
            df = pd.read_sql_table(table, con=engine)
            df.to_excel(writer, sheet_name=table, index=False)

    # Auto-adjust column widths
    wb = load_workbook(output_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            adjusted_width = max_length + 2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
    wb.save(output_file)

# -------------------------
# Run
# -------------------------
if __name__ == "__main__":
    generate_data(n_patients=20)
    export_to_excel()
